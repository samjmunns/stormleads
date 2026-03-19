"""
US Treasury FIO ZIP-level homeowners insurance data client.

Downloads the public Excel dataset from the US Treasury Federal Insurance Office (FIO)
once, parses it, and caches results to disk for 30 days.

Data: 2018–2022 ZIP-level homeowners insurance market metrics.
Source: US Dept of the Treasury — home.treasury.gov (FIO Annual Report Supporting Data)

Metrics returned per ZIP:
  policy_count  — number of homeowners insurance policies written
  avg_premium   — average annual premium ($)
  loss_ratio    — incurred losses / earned premium (lower = less risky to insure)

These are combined with Census housing unit counts in lead_scorer.py to compute
insurance penetration rate (policies per housing unit) per ZIP.

Note: If the Treasury FIO URL is unreachable or the Excel format changes,
this client falls back to empty data and the lead scorer uses neutral scores.
"""
import io
import json
import logging
from pathlib import Path

import httpx

logger = logging.getLogger(__name__)

TREASURY_URL = (
    "https://home.treasury.gov/system/files/311/"
    "Supporting_Underlying_Metrics_and_Disclaimer_for_Analyses_of_US_"
    "Homeowners_Insurance_Markets_2018-2022.xlsx"
)

CACHE_FILE = Path("data/insurance_cache.json")
HEADERS = {"User-Agent": "StormLeads/1.0 (contact@stormleads.com)"}

# Module-level cache so we only parse once per server restart
_cache: dict | None = None


async def get_insurance_data(zip_codes: list[str]) -> dict[str, dict]:
    """
    Returns insurance metrics for requested ZIP codes.

    Each entry contains:
      policy_count  — raw policy count (2018–2022 avg)
      avg_premium   — average annual premium ($)
      loss_ratio    — incurred losses / earned premium
    Falls back to empty dict if data is unavailable.
    """
    global _cache
    if _cache is None:
        _cache = await _load_or_fetch()
    return {z: _cache[z] for z in zip_codes if z in _cache}


async def _load_or_fetch() -> dict:
    """Load from disk cache or download + parse from Treasury."""
    if CACHE_FILE.exists():
        try:
            data = json.loads(CACHE_FILE.read_text())
            logger.info(f"Insurance cache: loaded {len(data)} ZIPs from disk")
            return data
        except Exception as e:
            logger.warning(f"Insurance cache read failed ({e}), re-fetching")
    return await _fetch_and_parse()


async def _fetch_and_parse() -> dict:
    """Download Treasury FIO Excel and extract ZIP-level insurance data."""
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        logger.warning("openpyxl not installed — insurance data unavailable. Run: pip install openpyxl")
        return {}

    logger.info("Treasury FIO: downloading homeowners insurance Excel (one-time)...")
    try:
        async with httpx.AsyncClient(headers=HEADERS, timeout=120.0) as client:
            resp = await client.get(TREASURY_URL)
            resp.raise_for_status()
            excel_bytes = resp.content
    except httpx.HTTPError as e:
        logger.warning(f"Treasury FIO download failed: {e} — insurance scoring disabled")
        return {}

    logger.info(f"Treasury FIO: downloaded {len(excel_bytes):,} bytes, parsing...")
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), read_only=True, data_only=True)
        results = _parse_workbook(wb)
    except Exception as e:
        logger.warning(f"Treasury FIO parse failed: {e}")
        return {}

    if results:
        CACHE_FILE.parent.mkdir(exist_ok=True)
        CACHE_FILE.write_text(json.dumps(results))
        logger.info(f"Treasury FIO: parsed {len(results)} ZIPs, cached to disk")
    else:
        logger.warning("Treasury FIO: no ZIP data found in workbook")

    return results


def _parse_workbook(wb) -> dict:
    """
    Walk sheets looking for ZIP-level data with policy/premium/loss columns.
    Returns dict keyed by 5-digit ZIP string.
    """
    # Column name patterns to look for (case-insensitive substrings)
    ZIP_PATTERNS = ("zip", "zcta", "postal code")
    POLICY_PATTERNS = ("polic", "count", "number of")
    PREMIUM_PATTERNS = ("premium", "written prem", "earned prem")
    LOSS_PATTERNS = ("incurred loss", "loss incurred", "total loss")

    def find_col(headers: list[str], patterns: tuple) -> int | None:
        for pat in patterns:
            for i, h in enumerate(headers):
                if pat in h:
                    return i
        return None

    def safe_float(val) -> float:
        try:
            return float(val) if val is not None else 0.0
        except (TypeError, ValueError):
            return 0.0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 10:
            continue

        # Find header row (first row with a ZIP-like column)
        header_row_idx = None
        for ri, row in enumerate(rows[:10]):
            normalized = [str(c).lower().strip() if c else "" for c in row]
            if find_col(normalized, ZIP_PATTERNS) is not None:
                header_row_idx = ri
                break

        if header_row_idx is None:
            continue

        headers = [str(c).lower().strip() if c else "" for c in rows[header_row_idx]]
        zip_col = find_col(headers, ZIP_PATTERNS)
        policy_col = find_col(headers, POLICY_PATTERNS)
        premium_col = find_col(headers, PREMIUM_PATTERNS)
        loss_col = find_col(headers, LOSS_PATTERNS)

        if zip_col is None:
            continue

        logger.info(
            f"Treasury FIO: parsing sheet '{sheet_name}' — "
            f"zip={zip_col} policy={policy_col} premium={premium_col} loss={loss_col}"
        )

        results = {}
        for row in rows[header_row_idx + 1:]:
            if not row or len(row) <= zip_col:
                continue
            raw_zip = str(row[zip_col] or "").strip()
            # Normalize to 5-digit zero-padded ZIP
            z = raw_zip.zfill(5) if raw_zip.isdigit() else ""
            if not z or len(z) != 5:
                continue

            policies = safe_float(row[policy_col] if policy_col is not None else None)
            premium = safe_float(row[premium_col] if premium_col is not None else None)
            losses = safe_float(row[loss_col] if loss_col is not None else None)

            results[z] = {
                "policy_count": round(policies),
                "avg_premium": round(premium / policies) if policies > 0 else 0,
                "loss_ratio": round(losses / premium, 3) if premium > 0 else 0.0,
            }

        if results:
            return results

    return {}
